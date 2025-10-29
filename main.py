import os
import re
import json
import pandas as pd
from transformers import AutoTokenizer
from ibm_watsonx_ai import APIClient
from ibm_watsonx_ai.metanames import EmbedTextParamsMetaNames
from langchain_ibm import WatsonxEmbeddings, WatsonxLLM
from pymilvus import MilvusClient
from ibm_watsonx_ai.metanames import GenTextParamsMetaNames as GenParams
from sklearn.preprocessing import normalize 
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()
my_credentials = {
    "url": os.getenv("URL"),
    "apikey": os.getenv("APIKEY"),
}
project_id = os.getenv("PROJECT_ID")
model_id = "intfloat/multilingual-e5-large"

# トークナイザー初期化 
tokenizer = AutoTokenizer.from_pretrained(model_id)

# テキスト分割 
def chunk_texts(texts, max_tokens=510):
    chunks = []
    for text in texts:
        tokens = tokenizer.encode(text, add_special_tokens=False)
        for i in range(0, len(tokens), max_tokens):
            chunk_tokens = tokens[i:i + max_tokens]
            chunk_text = tokenizer.decode(chunk_tokens)
            token_count = len(tokenizer.encode(chunk_text, add_special_tokens=True))
            if token_count <= 512:
                chunks.append(chunk_text)
    return chunks

# ファイル読み込みと分割
def load_and_chunk_file(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        raw_text = f.read()
    return chunk_texts([raw_text])

# 埋め込み生成 
def get_embeddings(texts):
    embedder = WatsonxEmbeddings(
        model_id=model_id,
        url=my_credentials["url"],
        apikey=my_credentials["apikey"],
        project_id=project_id
    )
    vectors = embedder.embed_documents(texts=texts)
    return normalize(vectors).tolist()

# Milvusセットアップ 
def setup_milvus(vectors, texts, db_path="./milvus_yuhou.db", collection="yuhou_collection"):
    milvus = MilvusClient(db_path)
    if milvus.has_collection(collection):
        milvus.drop_collection(collection)
    milvus.create_collection(
        collection_name=collection,
        dimension=len(vectors[0]),
        metric_type="IP",
        consistency_level="Bounded"
    )
    data = [{"id": i, "vector": vec, "text": texts[i]} for i, vec in enumerate(vectors)]
    milvus.insert(collection_name=collection, data=data)
    return milvus

# 出力処理
def export_markdown(query, context, response, output_dir="outputs"):
    os.makedirs(output_dir, exist_ok=True)
    with open(os.path.join(output_dir, "results.md"), "a", encoding="utf-8") as f:
        f.write(f"## クエリ: {query}\n\n### 根拠:\n{context}\n\n### 回答:\n{response}\n\n---\n\n")

def export_json(query, context, response, output_file="outputs/results.json"):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    data = []
    if os.path.exists(output_file):
        with open(output_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    data.append({"query": query, "context": context, "response": response})
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def export_to_excel(results, output_file="outputs/results.xlsx"):
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.border = border
    wb.save(output_file)


file_path = "data/nissan.plain.txt" 
texts = load_and_chunk_file(file_path)
embedding_vectors = get_embeddings(texts)
milvus = setup_milvus(embedding_vectors, texts)

queries = [
    "退職給付関係に基づいて、退職金を教えてください。"
]

query_vectors = get_embeddings(queries)

# LLM設定
generate_params = {
    GenParams.MAX_NEW_TOKENS: 500,
    GenParams.MIN_NEW_TOKENS: 0,
    GenParams.DECODING_METHOD: "greedy",
    GenParams.REPETITION_PENALTY: 1
}

custom_llm = WatsonxLLM(
    model_id="meta-llama/llama-4-maverick-17b-128e-instruct-fp8",
    url=my_credentials["url"],
    apikey=my_credentials["apikey"],
    project_id=project_id,
    params=generate_params,
)

# クエリ回答 
all_results = []

for query, query_vector in zip(queries, query_vectors):
    search_results = milvus.search(
        collection_name="yuhou_collection",
        data=[query_vector],
        limit=10,
        output_fields=["id", "text"]
    )

    hits = search_results[0]
    context = "\n\n".join([
        f"根拠 {i+1}（ID: {hit['id']}）\n" + re.sub(r"\s{2,}", " ", hit["text"].strip())
        for i, hit in enumerate(hits)
    ])

    prompt = f"""
    ## System
    あなたは財務分析に精通したアシスタントです。
    ユーザーの指示に可能な限り正確に従ってください。
    知ったかぶりをせず、事実に基づいて回答してください。
    必ず日本語で回答してください。

    ## User
    質問:
    {query}

    根拠:
    {context}

    回答:
    以下の手順に従って、質問に対する財務分析を行ってください。

    1. 質問文から主要キーワードを抽出してください（例：「退職」「未払金」「買掛金」など）。
    2. 抽出したキーワードの同義語・類義語も列挙してください。
    - 例：「退職給付費用」「退職一時金」は「退職金」と同義として扱うこと。
    3. 根拠の中から、抽出したキーワードおよび同義語を含むすべての記述（表の該当行や文章）を抜き出してください。
    4. 抽出した根拠に基づき、該当する数値や説明を簡潔にまとめてください。
    5. 数値の計算が必要な場合は、根拠に記載された算出方法に従って計算し、計算のステップ（式や途中結果）と最終的な計算結果を示してください。
    - 例：「退職金」の場合、引当金がある場合は、積み立てた引当金から支払う；引当金がない場合は、支払い時に初めて費用を計上する
    6. 回答は日本語で、簡潔かつ正確に記述してください。
    
    【重要】
    ・回答は途中で切らず、最後まで出力してください。


    """

    response = custom_llm.invoke(prompt)

    print(f"=== クエリ: {query} ===")
    print(response)
    print("\n" + "="*50 + "\n")

    export_markdown(query, context, response)
    export_json(query, context, response)

    all_results.append({"質問": query, "根拠": context, "回答": response})

export_to_excel(all_results)
